import openpyxl

# Load the Excel file
workbook = openpyxl.load_workbook('Instrumentos 2019-2024_19_11.xlsx')

# Select the worksheet named "Empenhos"
worksheet = workbook['Empenhos']

# Create a new column header for the links
worksheet.cell(row=1, column=6, value='Links')

# Iterate over the rows in the worksheet starting from the second row
for row in range(2, worksheet.max_row + 1):
    # Get the formula in column C
    formula = worksheet.cell(row=row, column=3).value
    
    # Extract the link from the formula
    if formula and formula.startswith('=HYPERLINK('):
        link = formula.split('"')[1]
    else:
        link = None
    
    # Write the link to the new column (column F)
    worksheet.cell(row=row, column=6, value=link)

# Save the updated workbook
workbook.save('empenhos_updated.xlsx')

print("Links have been extracted and saved to a new column in 'empenhos_updated.xlsx'.")
